"""
stock_entry_import_service.py — Excel orqali omborga KIRIM qilish.

Oqim:
    Excel fayl  →  parse + validatsiya  →  mahsulotlarni aniqlash (barcode/sku/nom)
              →  items ro'yxati  →  StockEntryService.create_entry()  →  StockEntry

Supplier, cash_amount, card_amount va do'kon API darajasida beriladi;
mahsulot satrlari (nom/barcode/sku, miqdor, narxlar) Excelda joylashadi.

Bazada topilmagan mahsulotlar bilan ishlash:
    - analyze_from_excel()  — faylni import qilmasdan tahlil qiladi, yangi
      mahsulot nomzodlarini "new_products" sifatida qaytaradi (frontend
      foydalanuvchidan qo'shish yoki tashlab ketishni so'raydi);
    - import_from_excel(create_products=True)  — nomzodlarni Product sifatida
      yaratib, kirimga qo'shadi (hammasi bitta tranzaksiyada);
    - import_from_excel(create_products=False) — nomzod satrlar o'tkazib
      yuboriladi, faqat mavjud mahsulotlar kirim qilinadi.
"""
from decimal import Decimal, InvalidOperation

import openpyxl
from django.core.exceptions import ValidationError
from django.db import transaction

from apps.products.models import Product
from apps.products.utils.barcode_utility import normalize_barcode
from apps.contract.services.stock_entry_service import StockEntryService


REQUIRED_COLUMNS = {"quantity", "purchase_price", "selling_price"}
IDENTIFIER_COLUMNS = {"barcode", "sku", "name"}

HEADER_MAP = {
    # Identifikatorlar
    "barcode":          "barcode",
    "shtrix kod":       "barcode",
    "shtrix-kod":       "barcode",
    "sku":              "sku",
    "artikul":          "sku",
    "name":             "name",
    "nomi":             "name",
    "mahsulot":         "name",
    "mahsulot nomi":    "name",
    # Miqdor
    "quantity":         "quantity",
    "miqdor":           "quantity",
    "miqdori":          "quantity",
    "soni":             "quantity",
    # Narxlar
    "purchase_price":     "purchase_price",
    "sotib olingan narx": "purchase_price",
    "kirim narxi":        "purchase_price",
    "xarid narxi":        "purchase_price",
    "tannarx":            "purchase_price",
    "selling_price":      "selling_price",
    "sotish narxi":       "selling_price",
    "sotuv narxi":        "selling_price",
    "wholesale_price":    "wholesale_price",
    "optom narx":         "wholesale_price",
    "ulgurji narx":       "wholesale_price",
}


class StockEntryImportService:

    # Yangi mahsulot yaratishda Product modelidagi CharField cheklovlari
    MAX_NAME_LEN = 100
    MAX_SKU_LEN = 64

    @classmethod
    def analyze_from_excel(cls, *, file) -> dict:
        """
        Excel faylni import qilmasdan tahlil qiladi.

        Qaytaradi:
            {
              "ready": int,              # mavjud mahsulotlarga mos kelgan satrlar soni
              "new_products": [          # bazada topilmagan, yaratsa bo'ladigan satrlar
                {"row", "name", "barcode", "sku", "quantity",
                 "purchase_price", "selling_price", "wholesale_price"}
              ],
              "skipped": [ {"row", "reason"} ],
            }
        """
        col, data_rows = cls._read_sheet(file)
        parsed, skipped = cls._parse_rows(col, data_rows)
        product_maps = cls._build_product_maps(parsed)
        resolved, candidates, split_skipped = cls._split_rows(parsed, product_maps)
        skipped += split_skipped

        return {
            "ready": len(resolved),
            "new_products": [
                {
                    "row": c["row"],
                    "name": c["name"],
                    "barcode": c["barcode"],
                    "sku": c["sku"],
                    "quantity": c["quantity"],
                    "purchase_price": str(c["purchase_price"]),
                    "selling_price": str(c["selling_price"]),
                    "wholesale_price": str(c["wholesale_price"]),
                }
                for c in candidates
            ],
            "skipped": skipped,
        }

    @classmethod
    def import_from_excel(cls, *, file, supplier, store, cash_amount, card_amount, user,
                          create_products=False) -> dict:
        """
        Excel fayldan kirim yaratadi.

        create_products=True bo'lsa bazada topilmagan mahsulotlar avval Product
        sifatida yaratiladi (kirim bilan bitta tranzaksiyada), aks holda bunday
        satrlar o'tkazib yuboriladi.

        Qaytaradi:
            {
              "entry_id": int | None,   # yaratilgan kirim (valid satr bo'lmasa None)
              "created": int,           # kirimga qo'shilgan satrlar soni
              "created_products": int,  # yangi yaratilgan mahsulotlar soni
              "skipped": [ {"row", "reason"} ],  # o'tkazib yuborilgan satrlar
              "total_amount": str,
              "paid_amount": str,
              "debt_amount": str,
              "payment_type": str | None,
            }
        """
        col, data_rows = cls._read_sheet(file)
        parsed, skipped = cls._parse_rows(col, data_rows)
        product_maps = cls._build_product_maps(parsed)
        resolved, candidates, split_skipped = cls._split_rows(parsed, product_maps)
        skipped += split_skipped

        created_products = 0
        # Mahsulot yaratish + kirim bitta tranzaksiyada: kirim yaratilmasa
        # (masalan to'lov validatsiyasi yiqilsa) yangi mahsulotlar ham qolmaydi.
        with transaction.atomic():
            if candidates:
                if create_products:
                    created, create_skipped = cls._create_products(candidates)
                    skipped += create_skipped
                    created_products = len(created)
                    resolved += created
                    resolved.sort(key=lambda pair: pair[0]["row"])  # Excel satr tartibini saqlaymiz
                else:
                    skipped += [
                        {
                            "row": c["row"],
                            "reason": "mahsulot bazada topilmadi (yangi mahsulot qo'shish tanlanmadi): "
                                      f"{c['name'] or c['barcode'] or c['sku']}",
                        }
                        for c in candidates
                    ]

            items = [
                {
                    "product": product,
                    "quantity": p["quantity"],
                    "purchase_price": p["purchase_price"],
                    "selling_price": p["selling_price"],
                    "wholesale_price": p["wholesale_price"],
                }
                for p, product in resolved
            ]

            if not items:
                return {
                    "entry_id": None,
                    "created": 0,
                    "created_products": 0,
                    "skipped": skipped,
                    "total_amount": "0.00",
                    "paid_amount": "0.00",
                    "debt_amount": "0.00",
                    "payment_type": None,
                }

            # To'lov validatsiyasi (faqat valid satrlar bo'yicha); atomic ichida
            # raise — yaratilgan mahsulotlar ham rollback bo'ladi
            total_amount = sum((i["purchase_price"] * i["quantity"] for i in items), Decimal("0"))
            paid_amount = Decimal(cash_amount) + Decimal(card_amount)
            if paid_amount > total_amount:
                raise ValidationError(
                    f"To'lov ({paid_amount}) umumiy kirim summasidan ({total_amount}) oshib ketdi."
                )

            entry = StockEntryService.create_entry(
                supplier=supplier,
                store=store,
                items=items,
                cash_amount=cash_amount,
                card_amount=card_amount,
                user=user,
            )

        return {
            "entry_id": entry.id,
            "created": len(items),
            "created_products": created_products,
            "skipped": skipped,
            "total_amount": str(entry.total_amount),
            "paid_amount": str(entry.paid_amount),
            "debt_amount": str(entry.debt_amount),
            "payment_type": entry.payment_type,
        }

    # ── Excel o'qish ──────────────────────────────────────────────────────────

    @classmethod
    def _read_sheet(cls, file):
        try:
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        except Exception:
            raise ValidationError("Excel faylni o'qib bo'lmadi. Fayl .xlsx formatida bo'lishi kerak.")

        ws = wb.active
        if ws is None:
            raise ValidationError("Excel faylda ishchi varaq topilmadi.")

        # YAXSHI: read_only=True + data_only=True - openpyxl formulalarni hisoblamaydi va lazy o'qiydi.
        # MUAMMO [KRITIK]: list(ws.iter_rows(...)) BARCHA satrlarni bir zumda xotiraga materializatsiya qiladi.
        #   ~48k qatorli faylda bu katta RAM va sekin request demak. iter_rows generatorining afzalligi yo'qoladi.
        # YECHIM: satr sonini oldindan cheklash yoki chunk-lab ishlash:
        #   MAX_ROWS = 5000; ws.max_row tekshiruvi yoki enumerate(ws.iter_rows(...)) ni generator sifatida uzatish.
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValidationError("Excel fayl bo'sh.")

        # Sarlavhalarni normallashtiramiz: majburiylik belgisi "*" va ortiqcha bo'shliqlar olib tashlanadi.
        raw_headers    = [str(h).strip().lower().rstrip("*").strip() if h is not None else "" for h in rows[0]]
        mapped_headers = [HEADER_MAP.get(h, h) for h in raw_headers]
        col = {name: idx for idx, name in enumerate(mapped_headers)}

        missing = REQUIRED_COLUMNS - set(mapped_headers)
        if missing:
            raise ValidationError(f"Ustunlar topilmadi: {', '.join(sorted(missing))}")

        if not (IDENTIFIER_COLUMNS & set(mapped_headers)):
            raise ValidationError(
                "Kamida bitta mahsulot identifikatori ustuni kerak: 'Barcode', 'SKU' yoki 'Mahsulot nomi'."
            )

        data_rows = rows[1:]
        if not data_rows:
            raise ValidationError("Shablon bo'sh — ma'lumot qatorlari yo'q.")

        return col, data_rows

    # ── Satrlarni parse qilish ─────────────────────────────────────────────────

    @classmethod
    def _parse_rows(cls, col, data_rows):
        """
        Excel satrlarini parse + validatsiya qiladi.
        Qaytaradi: (parsed, skipped)
            parsed  — [{"row", "barcode", "sku", "name", "quantity",
                        "purchase_price", "selling_price", "wholesale_price"}]
            skipped — [{"row", "reason"}]
        """
        parsed  = []
        skipped = []

        has = {key: (key in col) for key in IDENTIFIER_COLUMNS}

        for row_num, row in enumerate(data_rows, start=2):
            barcode = cls._cell(row, col["barcode"]) if has["barcode"] else ""
            sku     = cls._cell(row, col["sku"])     if has["sku"]     else ""
            name    = cls._cell(row, col["name"])    if has["name"]    else ""

            # To'liq bo'sh satrni jimgina o'tkazamiz
            if not any([barcode, sku, name]) and cls._row_is_empty(row, col):
                continue

            if not any([barcode, sku, name]):
                skipped.append({"row": row_num, "reason": "mahsulot identifikatori yo'q (barcode/sku/nom)"})
                continue

            quantity, err = cls._parse_quantity(cls._cell(row, col["quantity"]))
            if err:
                skipped.append({"row": row_num, "reason": f"miqdor: {err}"})
                continue

            purchase_price, err = cls._parse_decimal(cls._cell(row, col["purchase_price"]))
            if err:
                skipped.append({"row": row_num, "reason": f"sotib olingan narx: {err}"})
                continue

            selling_price, err = cls._parse_decimal(cls._cell(row, col["selling_price"]))
            if err:
                skipped.append({"row": row_num, "reason": f"sotish narxi: {err}"})
                continue

            if "wholesale_price" in col:
                wholesale_price, err = cls._parse_decimal(cls._cell(row, col["wholesale_price"]), default=Decimal("0"))
                if err:
                    skipped.append({"row": row_num, "reason": f"optom narx: {err}"})
                    continue
            else:
                wholesale_price = Decimal("0")

            reason = cls._validate_business(quantity, purchase_price, selling_price, wholesale_price)
            if reason:
                skipped.append({"row": row_num, "reason": reason})
                continue

            parsed.append({
                "row": row_num,
                "barcode": barcode,
                "sku": sku,
                "name": name,
                "quantity": quantity,
                "purchase_price": purchase_price,
                "selling_price": selling_price,
                "wholesale_price": wholesale_price,
            })

        return parsed, skipped

    # ── Mahsulotlarni aniqlash ─────────────────────────────────────────────────

    @classmethod
    def _build_product_maps(cls, parsed) -> dict:
        """barcode/sku/nom bo'yicha bitta-bittadan query — faqat ACTIVE mahsulotlar."""
        # YAXSHI: Mahsulotlar loop ichida emas, __in orqali 3 ta jamlangan query bilan olinadi (barcode/sku/name).
        #   Bu N+1 ni oldini oladi - _resolve_product keyin faqat xotiradagi map'dan o'qiydi. Namunali yechim.
        # Eslatma [PERF]: name uchun annotate(Lower("name")) filtri indeksdan foydalanmaydi (funktsional indeks yo'q),
        #   agar Product jadvali katta bo'lsa nom bo'yicha qidiruv sekin bo'lishi mumkin. barcode/sku afzal.
        # Barcode xom holda ham, EAN-13 normallashtirilgan holda ham qidiriladi —
        # Excelda 12 raqamli (checksumsiz) yozilgan barcode mavjud mahsulotga mos kelsin.
        barcodes = set()
        for p in parsed:
            if p["barcode"]:
                barcodes.add(p["barcode"])
                norm = cls._try_normalize_barcode(p["barcode"])
                if norm:
                    barcodes.add(norm)
        skus     = {p["sku"] for p in parsed if p["sku"]}
        names    = {p["name"].lower() for p in parsed if p["name"]}

        active = Product.objects.filter(status=Product.ProductStatus.ACTIVE)

        barcode_map = {}
        if barcodes:
            barcode_map = {
                p.barcode: p
                for p in active.filter(barcode__in=barcodes)
            }

        sku_map = {}
        if skus:
            sku_map = {
                p.sku: p
                for p in active.filter(sku__in=skus)
            }

        # Nom case-insensitive; nom UNIKAL emas — bir nechta mos kelsa ambiguity belgilanadi.
        name_map = {}
        if names:
            from django.db.models.functions import Lower
            qs = active.annotate(_lname=Lower("name")).filter(_lname__in=names)
            for p in qs:
                key = p.name.lower()
                if key in name_map:
                    name_map[key] = "AMBIGUOUS"
                elif name_map.get(key) != "AMBIGUOUS":
                    name_map[key] = p

        return {"barcode": barcode_map, "sku": sku_map, "name": name_map}

    @classmethod
    def _split_rows(cls, parsed, maps):
        """
        Parse qilingan satrlarni uch guruhga ajratadi:
            resolved   — [(satr, Product)] mavjud faol mahsulotga mos kelganlar
            candidates — [satr] bazada topilmagan, yangi mahsulot sifatida
                         yaratsa bo'ladiganlar (nomi bor, cheklovlarga sig'adi)
            skipped    — [{"row", "reason"}] xato/yaramaydigan satrlar
        """
        resolved, candidates, skipped = [], [], []
        seen_pids = set()
        seen_new  = set()  # nomzod identifikatorlari — fayl ichidagi dublikatlar uchun

        for p in parsed:
            product, reason = cls._resolve_product(p, maps)

            if product is not None:
                # Bitta kirimda bir mahsulot ikki marta kelsa — dublikat batch oldini olamiz
                if product.id in seen_pids:
                    skipped.append({"row": p["row"], "reason": f"mahsulot satrlarda takrorlangan: {product.name}"})
                    continue
                seen_pids.add(product.id)
                resolved.append((p, product))
                continue

            if reason is not None:
                # Aniq xato (masalan nom bir nechta mahsulotga mos) — nomzod emas
                skipped.append({"row": p["row"], "reason": reason})
                continue

            # Bazada topilmadi — yangi mahsulot nomzodi
            creation_error = cls._validate_new_product(p)
            if creation_error:
                skipped.append({"row": p["row"], "reason": creation_error})
                continue

            # Yaratish va dublikat tekshiruvi yagona formatda bo'lishi uchun
            # barcode EAN-13 ga normallashtiriladi (12 raqam → checksum bilan 13)
            if p["barcode"]:
                p["barcode"] = cls._try_normalize_barcode(p["barcode"])

            keys = {k for k in (p["barcode"], p["sku"], p["name"].lower()) if k}
            if keys & seen_new:
                skipped.append({"row": p["row"], "reason": f"yangi mahsulot satrlarda takrorlangan: {p['name']}"})
                continue
            seen_new |= keys
            candidates.append(p)

        return resolved, candidates, skipped

    @classmethod
    def _resolve_product(cls, p, maps):
        """
        Ustuvorlik: barcode → sku → nom.
        Qaytaradi:
            (product, None) — mavjud faol mahsulot topildi
            (None, sabab)   — aniq xato (nom bir nechta mahsulotga mos keldi)
            (None, None)    — bazada topilmadi (yangi mahsulot nomzodi)
        """
        if p["barcode"]:
            product = maps["barcode"].get(p["barcode"])
            if product is None:
                # Xom qiymat topilmasa normallashtirilgan EAN-13 bilan urinamiz
                norm = cls._try_normalize_barcode(p["barcode"])
                if norm:
                    product = maps["barcode"].get(norm)
            return product, None

        if p["sku"]:
            return maps["sku"].get(p["sku"]), None

        match = maps["name"].get(p["name"].lower())
        if match == "AMBIGUOUS":
            return None, f"nom bir nechta mahsulotga mos keldi (barcode/sku ishlating): {p['name']}"
        return match, None

    # ── Yangi mahsulotlarni yaratish ───────────────────────────────────────────

    @classmethod
    def _validate_new_product(cls, p):
        """Satr yangi mahsulot sifatida yaratishga yaroqlimi. None yoki sabab qaytaradi."""
        if not p["name"]:
            ident = p["barcode"] or p["sku"]
            return f"mahsulot topilmadi, yangi mahsulot sifatida qo'shish uchun nom ustuni kerak: {ident}"
        if len(p["name"]) > cls.MAX_NAME_LEN:
            return f"mahsulot nomi {cls.MAX_NAME_LEN} belgidan oshmasligi kerak: {p['name'][:30]}…"
        # Product.save() barcode uchun EAN-13 shtrix rasm generatsiya qiladi —
        # yaroqsiz barcode bilan yaratish mumkin emas
        if p["barcode"] and cls._try_normalize_barcode(p["barcode"]) is None:
            return f"yangi mahsulot uchun barcode yaroqli EAN-13 bo'lishi kerak (12-13 raqam): {p['barcode']}"
        if p["sku"] and len(p["sku"]) > cls.MAX_SKU_LEN:
            return f"sku {cls.MAX_SKU_LEN} belgidan oshmasligi kerak: {p['sku']}"
        return None

    @staticmethod
    def _try_normalize_barcode(value):
        """Barcode'ni EAN-13 formatga keltiradi; yaroqsiz bo'lsa None qaytaradi."""
        try:
            return normalize_barcode(value)
        except ValueError:
            return None

    @classmethod
    def _create_products(cls, candidates):
        """
        Nomzod satrlardan yangi Product yozuvlarini yaratadi.

        barcode/sku UNIQUE — noaktiv/draft mahsulotlar bilan to'qnashuv bo'lishi
        mumkin (maps faqat ACTIVE dan qurilgan), shuning uchun barcha statuslar
        bo'yicha oldindan tekshiramiz.

        Qaytaradi: ([(satr, Product)], skipped)
        """
        barcodes = {c["barcode"] for c in candidates if c["barcode"]}
        skus     = {c["sku"] for c in candidates if c["sku"]}
        taken_barcodes = (
            set(Product.objects.filter(barcode__in=barcodes).values_list("barcode", flat=True))
            if barcodes else set()
        )
        taken_skus = (
            set(Product.objects.filter(sku__in=skus).values_list("sku", flat=True))
            if skus else set()
        )

        created, skipped = [], []
        for c in candidates:
            if c["barcode"] and c["barcode"] in taken_barcodes:
                skipped.append({
                    "row": c["row"],
                    "reason": f"barcode faol bo'lmagan mahsulotga tegishli: {c['barcode']}",
                })
                continue
            if c["sku"] and c["sku"] in taken_skus:
                skipped.append({
                    "row": c["row"],
                    "reason": f"sku faol bo'lmagan mahsulotga tegishli: {c['sku']}",
                })
                continue

            # sku/barcode berilmasa Product.save() ichida avtomatik generatsiya
            # qilinadi; name_uz aniq to'ldiriladi (modeltranslation, default 'uz').
            product = Product(
                name=c["name"],
                name_uz=c["name"],
                barcode=c["barcode"] or None,
                sku=c["sku"] or None,
            )
            product.save()
            created.append((c, product))

        return created, skipped

    # ── Yordamchi metodlar ─────────────────────────────────────────────────────

    @staticmethod
    def _cell(row, idx: int) -> str:
        val = row[idx]
        return str(val).strip() if val is not None else ""

    @staticmethod
    def _row_is_empty(row, col) -> bool:
        return all(StockEntryImportService._cell(row, idx) == "" for idx in col.values())

    @staticmethod
    def _parse_quantity(value: str):
        """Miqdor: kasr (0.25 qadam — juft mahsulot) qabul qilinadi; int() bilan KESILMAYDI.
        Mahsulot juft emasligi (butun son talabi) StockEntryService.create_stock_entry da tekshiriladi."""
        if value == "":
            return None, "bo'sh"
        try:
            parsed = Decimal(str(value).replace(" ", "").replace(",", "."))
        except (InvalidOperation, ValueError, TypeError):
            return None, "raqam emas"
        if parsed <= 0:
            return None, "0 dan katta bo'lishi kerak"
        if (parsed * 4) % 1 != 0:
            return None, "0.25 ga karrali bo'lishi kerak"
        return parsed, None

    @staticmethod
    def _parse_decimal(value: str, default=None):
        if value == "":
            if default is not None:
                return default, None
            return None, "bo'sh"
        try:
            parsed = Decimal(value.replace(" ", "").replace(",", "."))
        except (InvalidOperation, ValueError, TypeError):
            return None, "raqam emas"
        if parsed < 0:
            return None, "manfiy bo'lmasligi kerak"
        return parsed, None

    @staticmethod
    def _validate_business(quantity, purchase_price, selling_price, wholesale_price):
        """StockEntryItemSerializer dagi qoidalar bilan bir xil."""
        if purchase_price <= 0:
            return "sotib olingan narx 0 dan katta bo'lishi kerak"
        if selling_price <= 0:
            return "sotish narxi 0 dan katta bo'lishi kerak"
        if selling_price < purchase_price:
            return "sotish narxi sotib olingan narxdan past bo'lmasligi kerak"
        if wholesale_price > 0:
            if wholesale_price < purchase_price:
                return "optom narx tannarxdan past bo'lmasligi kerak"
            if wholesale_price > selling_price:
                return "optom narx sotish narxidan yuqori bo'lmasligi kerak"
        return None
